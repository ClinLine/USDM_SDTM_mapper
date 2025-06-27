import jsonata
import json
import openpyxl
import pandas as pd

def get_ID(ID_string):
    if len(ID_string) <2: # if the ID string is None, return empty strings
        return "", ""
    else:
        o = 1 #letter it is looking at
        while ID_string[o] != ":" and o+1 < len(ID_string): #looking for the end of the ID
            o += 1
        if o == len(ID_string)-1: # if the ID is not found, return empty strings
            return "", ID_string            
        else:
            Id = ID_string[1:o-1] # extracting the ID from the string
            ID_less = ID_string[o+1:]
            while ID_less[-1]==" " or ID_less[-1]=="'": 
                ID_less=ID_less[:-1] # remove trailing blanks or quotes
                if len(ID_less) == 0: # if the ID_less is empty, return empty strings
                    return Id, ""
            while ID_less[0]==" " or ID_less[0]=="'": ID_less=ID_less[1:] # remove starting blanks or quotes
            return Id, ID_less

# general function(s)
def string_to_list(input, result):
    n = 0 #letter it is looking at
    while input[n] != "}": #looking for the end of the list
        if input[n-1:n+2] == ", '" or input[n] in ["{"]: # looking for the start of a new item in the list
            n += 1
            m = n
            while m+1 < len(input) and input[m+1] not in ("}") and input[m:m+3] not in ("', '"): # looking for the end of the item
                m += 1
            resX=input[n:m+1]
            if resX[-1]==",": resX=resX[:-1]
            result.append(resX) # appending the item to the list
            n = m + 1
        else: 
            n += 1

def Parse_jsonata(codeSnip,data):
        if codeSnip is None:
            result = " "
        else:
            try:
                expr = jsonata.Jsonata(codeSnip)
                result = expr.evaluate(data)  
            except:
                result = "Error in expression " + codeSnip
        if result is None: result = " "
        result= str(result)
        if result == "": result = " "
        if result == "{}": result = " "
        try:
            result0 = result.replace("’", " ")
        except:
            result0 = ""
        if result0 == "": result0= " "
        if result0[0] == "[": 
                result0 = result0[1:-1]
                result0 = result0.replace("}, {", ", ")
        return result0

def Create_TE(wb, JsonInput):
    ti_sheet = wb['TE']
    r = ti_sheet.max_row
    for i in range(2, ti_sheet.max_row + 1):
        j=i-1
        #swap the rows and columns in the TI sheet
        varName=ti_sheet.cell(row=i, column=1).value
        ti_sheet.cell(row=1, column=j).value = varName    
        if varName == "STUDYID":
            StudyIdCodeSnip = ti_sheet.cell(row=i, column=7).value
            StudyIDColumn = j
        if varName == "DOMAIN":
            DomainResult =  ti_sheet.cell(row=i, column=8).value
            DomainColumn = j
    for i in range(r, ti_sheet.max_column):
        for j in range(1, r):
            ti_sheet.cell(row=j, column=i).value = ""
    # create empty id array for checking value alignment in different columns
    id = []
    # Print the value in the first and seventh column of each row in the 'TS Parameters' sheet
    with open(JsonInput, 'r') as file:
        data=json.load (file)
        studyId=Parse_jsonata(codeSnip=StudyIdCodeSnip,data=data)         
        print("StudyId: ", studyId)
        
        for i in range(2, ti_sheet.max_row + 1):
            # Get all the mapping information from the TS Parameters sheet
            if i not in [StudyIDColumn+1, DomainColumn+1]:
                codeSnip = ti_sheet.cell(row=i, column=7).value
                result2=Parse_jsonata(codeSnip=codeSnip,data=data)
                x=1
                c = i - 1
                if result2 != " ":
                    if result2[0] == "{":  # check if the result is a list
                        result3 = []
                        string_to_list(result2, result3)  # convert the string to a list
                        # filling ts sheet if it is a list 
                        skip = 0
                        for j in range(0, len(result3)):
                            if len(id) < len(result3):  # if the ID list is empty, append the first ID
                                idcheck, result4 = get_ID(result3[j])  # extracting the ID from the string
                                id.append(idcheck)  # appending the ID to the list
                            else:
                                idcheck, result4 = get_ID(result3[j])  # extracting the ID from the string
                                if idcheck != id[j+skip]:
                                    x += 1
                                    skip += 1
                            x += 1
                            ti_sheet.cell(row=x, column=c).value = result4
                    else:
                        # filling ts sheet if it is not a list
                        idcheck, result2 = get_ID(result2)
                        x += 1
                        ti_sheet.cell(row=x, column=c).value = result2
                else:
                    if len(id)> 0:
                        for j in range(len(id)):
                            x += 1
                            ti_sheet.cell(row=x, column=c).value = " "
        if len(id)> 0: # add the standard variable values for each row.
            x=1
            for j in range(len(id)):
                x += 1
                # filling the STUDYID, DOMAIN and TIVERS columns
                ti_sheet.cell(row=x, column=StudyIDColumn).value = studyId
                ti_sheet.cell(row=x, column=DomainColumn).value = DomainResult  
      