import jsonata
import json
import openpyxl
import pandas as pd
import definition

def Create_TV(wb, JsonInput):
    ta_sheet = wb['TV']
    r = ta_sheet.max_row
    rem = []
    for i in range(2, ta_sheet.max_row + 1):
        j=i-1
        #swap the rows and columns in the TI sheet
        varName=ta_sheet.cell(row=i, column=1).value
        ta_sheet.cell(row=1, column=j).value = varName    
        if i == 2: # varName == "STUDYID":
            StudyIdCodeSnip = ta_sheet.cell(row=i, column=7).value
            StudyIDColumn = j
        elif i == 3: # varName == "DOMAIN":
            DomainResult =  ta_sheet.cell(row=i, column=8).value
            DomainColumn = j
        elif i == 4: # varName == "VISITNUM":
            VisitNumCodeSnip = ta_sheet.cell(row=i, column=7).value
            VisitNumColumn = j
        elif i == 5: # varName == "VISIT":
            VisitCodeSnip = ta_sheet.cell(row=i, column=7).value
            VisitDescriptionColumn = j
        elif i == 6: # varName == "VISITDY":
            VisitdyCodeSnip = ta_sheet.cell(row=i, column=7).value
            VisitdyColumn = j
        elif i == 7: # varName == "ARMCD":
            ArmcdCodeSnip = ta_sheet.cell(row=i, column=7).value
            ArmcdColumn = j
        elif i == 8: # varName == "ARM":
            ArmCodeSnip = ta_sheet.cell(row=i, column=7).value
            ArmColumn = j
        elif i == 9: # varName == "TVENRL":
            TVENRLCodeSnip = ta_sheet.cell(row=i, column=7).value
            TVENRLColumn = j
        elif i == 10: # varName == "TVSTRL":
            TVSTRLCodeSnip = ta_sheet.cell(row=i, column=7).value
            TVSTRLColumn = j
        else:
            rem.append(j)

    for i in range(r, ta_sheet.max_column):
        for j in range(1, r):
            ta_sheet.cell(row=j, column=i).value = ""
    # create empty id array for checking value alignment in different columns
    id = []
    # Print the value in the first and seventh column of each row in the 'TS Parameters' sheet
    with open(JsonInput, 'r') as file:
        data=json.load (file)
        x = 1
        studyId=definition.Parse_jsonata(codeSnip=StudyIdCodeSnip,data=data)
        VisitNum=definition.Parse_jsonata(codeSnip=VisitNumCodeSnip,data=data)         
        Visit=definition.Parse_jsonata(codeSnip=VisitCodeSnip,data=data)
        Visitdy=definition.Parse_jsonata(codeSnip=VisitdyCodeSnip,data=data)
        Armcd=definition.Parse_jsonata(codeSnip=ArmcdCodeSnip,data=data)
        Arm=definition.Parse_jsonata(codeSnip=ArmCodeSnip,data=data)
        TVENRL=definition.Parse_jsonata(codeSnip=TVENRLCodeSnip ,data=data)
        TVSTRL=definition.Parse_jsonata(codeSnip=TVSTRLCodeSnip ,data=data)
        resultVisitNum = []
        resultVisit = []
        resultVisitdy = []
        resultArmcd = []
        resultArm = []
        resultTVENRL = []
        resultTVSTRL = []
        if VisitNum[0] == "{":  # check if the result is a list
            definition.string_to_list(VisitNum, resultVisitNum)
        if Visit[0] == "{":  # check if the result is a list
            definition.string_to_list(Visit, resultVisit)  # convert the string to a list
        else:
            resultVisit.append(Visit)
        if Visitdy[0] == "{":  # check if the result is a list
            definition.string_to_list(Visitdy, resultVisitdy)  # convert the string to a list
        else:
            resultVisitdy.append(Visitdy)
        if Armcd[0] == "{":  # check if the result is a list
            definition.string_to_list(Armcd, resultArmcd)
        else:
            resultArmcd.append(Armcd)
        if Arm[0] == "{":  # check if the result is a list
            definition.string_to_list(Arm, resultArm)
        else:   
            resultArm.append(Arm)
        if TVENRL[0] == "{":  # check if the result is a list
            definition.string_to_list(TVENRL, resultTVENRL)
        else:
            resultTVENRL.append(TVENRL)
        if TVSTRL[0] == "{":  # check if the result is a list
            definition.string_to_list(TVSTRL, resultTVSTRL)
        else:
            resultTVSTRL.append(TVSTRL)
        
        resultvisit2 = []
        resultvisitdy2 = []
        resultarmcd2 = []
        resultarm2 = []
        resultTVENRL2 = []
        resultTVSTRL2 = []
        VisitEncounter = []
        NextVisit = []

        for item in resultVisitNum:
            VisitEncounterID, VisitEncounterLess = definition.get_ID(item)
            VisitEncounter.append(VisitEncounterID)  # appending the ID to the list
            NextVisit.append(VisitEncounterLess)  # appending the less ID to the list
        for item in resultVisit:
            VisitID, VisitLess = definition.get_ID(item)  # extracting the ID from the string
            resultvisit2.append(VisitLess)  # appending the ID to the list
        for item in resultVisitdy:
            VisitdyID, VisitdyLess = definition.get_ID(item)  # extracting the ID from the string
            resultvisitdy2.append(VisitdyLess)  # appending the ID to the list
        for item in resultArmcd:
            ArmcdID, ArmcdLess = definition.get_ID(item)  # extracting the ID from the string
            resultarmcd2.append(ArmcdLess)  # appending the ID to the list
        for item in resultArm:
            ArmID, ArmLess = definition.get_ID(item)  # extracting the ID from the string
            resultarm2.append(ArmLess)  # appending the ID to the list
        for item in resultTVENRL:
            TVENRLID, TVENRLLess = definition.get_ID(item)  # extracting the ID from the string
            resultTVENRL2.append(TVENRLLess)  # appending the ID to the list
        for item in resultTVSTRL:   
            TVSTRLID, TVSTRLLess = definition.get_ID(item)  # extracting the ID from the string
            resultTVSTRL2.append(TVSTRLLess)  # appending the ID to the list
        
        VisitEncounter3 = []
        resultVisit3 = []
        resultVisitdy3 = []
        resultTVENRL3 = []
        resultTVSTRL3 = []

        currentEncounter = None
        for n in range(len(VisitEncounter)):
            for i in range(len(VisitEncounter)):
                if currentEncounter is None:
                    if NextVisit[i] == "None":
                        VisitEncounter3.insert(0, VisitEncounter[i])
                        resultVisit3.insert(0, resultvisit2[i])
                        if len(resultvisitdy2) > i:
                            resultVisitdy3.insert(0, resultvisitdy2[i])
                        if len(resultTVENRL2) > i:
                            resultTVENRL3.insert(0, resultTVENRL2[i])
                        if len(resultTVSTRL2) > i:
                            resultTVSTRL3.insert(0, resultTVSTRL2[i])
                        currentEncounter = VisitEncounter[i]
                        break
                elif NextVisit[i] == currentEncounter:
                    VisitEncounter3.insert(0, VisitEncounter[i])
                    resultVisit3.insert(0, resultvisit2[i])
                    if len(resultvisitdy2) > i:
                        resultVisitdy3.insert(0, resultvisitdy2[i])
                    if len(resultTVENRL2) > i:
                        resultTVENRL3.insert(0, resultTVENRL2[i])
                    if len(resultTVSTRL2) > i:
                        resultTVSTRL3.insert(0, resultTVSTRL2[i])
                    currentEncounter = VisitEncounter[i]
                    break      
                
        for j in range(len(resultarmcd2)):
            for i in range(len(resultVisit3)):
                x += 1
                ta_sheet.cell(row=x, column=StudyIDColumn).value = studyId
                ta_sheet.cell(row=x, column=DomainColumn).value = DomainResult
                if len(resultVisitdy3) > i:
                    ta_sheet.cell(row=x, column=VisitdyColumn).value = resultVisitdy3[i]
                else:
                    ta_sheet.cell(row=x, column=VisitdyColumn).value = " "
                ta_sheet.cell(row=x, column=VisitNumColumn).value = i+1
                ta_sheet.cell(row=x, column=VisitDescriptionColumn).value = resultVisit3[i]
                ta_sheet.cell(row=x, column=ArmcdColumn).value = resultarmcd2[j]
                ta_sheet.cell(row=x, column=ArmColumn).value = resultarm2[j]
                if len(resultTVENRL3) > i:
                    ta_sheet.cell(row=x, column=TVENRLColumn).value = resultTVENRL3[i]
                else:
                    ta_sheet.cell(row=x, column=TVENRLColumn).value = " "
                if len(resultTVSTRL3) > i:
                    ta_sheet.cell(row=x, column=TVSTRLColumn).value = resultTVSTRL3[i]
                else:
                    ta_sheet.cell(row=x, column=TVSTRLColumn).value = " "

        for j in rem:
            for i in range(2, x + 1):
                ta_sheet.cell(row=i, column=j).value = ""
