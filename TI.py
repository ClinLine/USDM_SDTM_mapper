import jsonata
import json
import openpyxl
import pandas as pd
import definition

def Create_TI(wb, JsonInput):
    ti_sheet = wb['TI']
    overflowcolumn = ti_sheet.max_row-1
    for i in range(2, ti_sheet.max_row + 1):
        j=i-1
        #swap the rows and columns in the TI sheet
        varName=ti_sheet.cell(row=i, column=1).value
        ti_sheet.cell(row=1, column=j).value = varName    
        if varName == "STUDYID":
            StudyIdCodeSnip = ti_sheet.cell(row=i, column=7).value
            StudyIDColumn = j
        if varName == "DOMAIN":
            DomainResult =  ti_sheet.cell(row=i, column=8).value
            DomainColumn = j
        if varName == "TIVERS":
            VersionCodeSnip =  ti_sheet.cell(row=i, column=7).value
            VersionColumn = j
        if varName == "IETEST":
            TestColumn = j

    # create empty id array for checking value alignment in different columns
    id = []
    # Print the value in the first and seventh column of each row in the 'TS Parameters' sheet
    with open(JsonInput, 'r') as file:
        data=json.load (file)
        studyId=definition.Parse_jsonata(codeSnip=StudyIdCodeSnip,data=data)         
        print("TI StudyId: ", studyId)

        versionResult=definition.Parse_jsonata(codeSnip=VersionCodeSnip,data=data)
        
        for i in range(2, ti_sheet.max_row + 1):
            # Get all the mapping information from the TI sheet
            if i not in [StudyIDColumn+1, DomainColumn+1, VersionColumn+1]:
                codeSnip = ti_sheet.cell(row=i, column=7).value
                result2=definition.Parse_jsonata(codeSnip=codeSnip,data=data)
                x=1
                c = i - 1
                if result2 != " ":
                    if result2[0] == "{":  # check if the result is a list
                        result3 = []
                        definition.string_to_list(result2, result3)  # convert the string to a list
                        # filling ti sheet if it is a list 
                        skip = 0
                        for j in range(0, len(result3)):
                            if len(id) < len(result3):  # if the ID list is empty, append the first ID
                                idcheck, result4 = definition.get_ID(result3[j])  # extracting the ID from the string
                                id.append(idcheck)  # appending the ID to the list
                            else:
                                idcheck, result4 = definition.get_ID(result3[j])  # extracting the ID from the string
                                if idcheck != id[j+skip]:
                                    x += 1
                                    skip += 1
                            x += 1
                            breakoff = 0
                            # resolve tags in result4
                            if c==TestColumn: 
                                result4=definition.ResolveTag(result4,data)
                            while len(result4) > 200:
                                for k in range(200, 0, -1):
                                    if result4[k] == " ":
                                        if breakoff == 0:
                                            ti_sheet.cell(row=x, column=c).value = result4[0:k]
                                        else:
                                            ti_sheet.cell(row=x, column=overflowcolumn+breakoff).value = result4[0:k]
                                        result4 = result4[k+1:len(result4)]
                                        breakoff += 1
                                        break
                            if breakoff == 0:
                                ti_sheet.cell(row=x, column=c).value = result4
                            else:
                                ti_sheet.cell(row=x, column=overflowcolumn+breakoff).value = result4
                    else:
                        # filling ts sheet if it is not a list
                        idcheck, result2 = definition.get_ID(result2)
                        x += 1
                        breakoff = 0
                        while len(result2) > 200:
                            for k in range(200, 0, -1):
                                if result2[k] == " ":
                                    if breakoff == 0:
                                        ti_sheet.cell(row=x, column=c).value = result2[0:k]
                                    else:
                                        ti_sheet.cell(row=x, column=overflowcolumn+breakoff).value = result2[0:k]
                                    result2 = result2[k+1:len(result2)]
                                    breakoff += 1
                                    break
                        if breakoff == 0:
                            ti_sheet.cell(row=x, column=c).value = result2
                        else:
                            ti_sheet.cell(row=x, column=overflowcolumn+breakoff).value = result2
                else:
                    if len(id)> 0:
                        for j in range(len(id)):
                            x += 1
                            ti_sheet.cell(row=x, column=c).value = " "
        if len(id)> 0: # add the standard variable values for each row.
            x=1
            for j in range(len(id)):
                x += 1
                # filling the STUDYID, DOMAIN and TIVERS columns
                ti_sheet.cell(row=x, column=StudyIDColumn).value = studyId
                ti_sheet.cell(row=x, column=DomainColumn).value = DomainResult
                ti_sheet.cell(row=x, column=VersionColumn).value = versionResult    
        if ti_sheet.max_column > 8:   
            for i in range(9, ti_sheet.max_column):
                newname = f"IETEST{i-7}"
                ti_sheet.cell(row=1, column=i).value = newname