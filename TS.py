import jsonata
import json
import openpyxl
import definition

def Create_TS(wb, JsonInput):
    ts0_sheet = wb['TS']
    ts_sheet = wb['TS Parameters']
    DomainResult = " "
    StudyIdCodeSnip = " "
    for i in range(2, ts0_sheet.max_row + 1):
        j=i-1
        #swap the rows and columns in the TS sheet
        varName=ts0_sheet.cell(row=i, column=1).value
        ts0_sheet.cell(row=1, column=j).value = varName
        
        # Get standard row value information for TS
        if varName == "STUDYID":
            StudyIdCodeSnip = ts0_sheet.cell(row=i, column=7).value
        if varName == "DOMAIN":
            DomainResult =  ts0_sheet.cell(row=i, column=8).value

    with open(JsonInput, 'r') as file:
        data=json.load (file)
        # Get StudyId first and start with a row id for the TS sheet            
        try:
            expr = jsonata.Jsonata (StudyIdCodeSnip)
            studyId= expr.evaluate(data)
        except:
            studyId = "no data "
        x=1
        # Then continue with the mappings in the TS Parameters sheet
        ts_sheet.cell(row=1, column=7).value = "Mapping Results"
        for i in range(2, ts_sheet.max_row + 1):
            # Get all the mapping information from the TS Parameters sheet
            MapName = ts_sheet.cell(row=i, column=1).value
            MapCode = ts_sheet.cell(row=i, column=2).value
            nfValue = ts_sheet.cell(row=i, column=8).value
            codeSnip = ts_sheet.cell(row=i, column=7).value
            result2=definition.Parse_jsonata(codeSnip,data)
            codeSnip = ts_sheet.cell(row=i, column=9).value
            resultCd=definition.Parse_jsonata(codeSnip,data)
            codeSnip = ts_sheet.cell(row=i, column=10).value
            resultCdRef=definition.Parse_jsonata(codeSnip,data)
            codeSnip = ts_sheet.cell(row=i, column=11).value
            resultCdVer=definition.Parse_jsonata(codeSnip,data)
            codeSnip = ts_sheet.cell(row=i, column=7).value        
            codeSnipCd = ts_sheet.cell(row=i, column=9).value   
            codeSnipCdRef = ts_sheet.cell(row=i, column=10).value   
            codeSnipCdVer = ts_sheet.cell(row=i, column=11).value   
    
            # replace the apostrophes with spaces
            if result2 != " " or nfValue != " ":
                if result2[0] == "{":  # check if the result is a list
                    result3 = []
                    definition.string_to_list(result2, result3)  # convert the string to a list
                    if resultCd != " ":
                        resultCd2 = []
                        definition.string_to_list(resultCd, resultCd2)  # convert the string to a list
                        resultCdRef2 = []
                        definition.string_to_list(resultCdRef, resultCdRef2)  # convert the string to a list
                        resultCdVer2 = []
                        definition.string_to_list(resultCdVer, resultCdVer2)  # convert the string to
                    # filling ts sheet if it is a list 
                    for j in range(0, len(result3)):
                        x += 1
                        id, result4 = definition.get_ID(result3[j])  # extract the ID from the result
                        if j == 0:
                            base_id = id  # store the base ID for the first item
                        else:
                            if id == "":  # if the ID is empty, use the base ID
                                id = base_id
                        ts0_sheet.cell(row=x, column=1).value = " "
                        ts0_sheet.cell(row=x, column=1).value = studyId
                        ts0_sheet.cell(row=x, column=2).value = DomainResult   
                        ts0_sheet.cell(row=x, column=3).value = j + 1
                        ts0_sheet.cell(row=x, column=4).value = id
                        ts0_sheet.cell(row=x, column=5).value = MapCode    
                        ts0_sheet.cell(row=x, column=6).value = MapName                
                        ts0_sheet.cell(row=x, column=7).value = result4  
                        ts0_sheet.cell(row=x, column=8).value = " "   
                        if result2 == " ": ts0_sheet.cell(row=x, column=8).value = nfValue
                        if resultCd != " ":
                            id, resultcd4 = definition.get_ID(resultCd2[j])
                            id, resultCdRef4 = definition.get_ID(resultCdRef2[j]) 
                            id, resultCdVer4 = definition.get_ID(resultCdVer2[j]) 
                            ts0_sheet.cell(row=x, column=9).value = resultcd4                  
                            ts0_sheet.cell(row=x, column=10).value = resultCdRef4 
                            ts0_sheet.cell(row=x, column=11).value = resultCdVer4
                elif result2 != " ":  # if the result is not a list
                    # filling ts sheet if it is not a list
                    x += 1
                    if result2[0] == "{":
                        id, result2 = definition.get_ID(result2)  # extract the ID from the result
                    
                    ts0_sheet.cell(row=x, column=1).value = " "
                    ts0_sheet.cell(row=x, column=1).value = studyId
                    ts0_sheet.cell(row=x, column=2).value = DomainResult   
                    ts0_sheet.cell(row=x, column=3).value = " "
                    ts0_sheet.cell(row=x, column=4).value = " "
                    ts0_sheet.cell(row=x, column=5).value = MapCode    
                    ts0_sheet.cell(row=x, column=6).value = MapName                
                    ts0_sheet.cell(row=x, column=7).value = result2   
                    ts0_sheet.cell(row=x, column=8).value = " "   
                    if result2 == " ": ts0_sheet.cell(row=x, column=8).value = nfValue	
                    if resultCd != " ":
                        id, resultCd = definition.get_ID(resultCd)
                        id, resultCdRef = definition.get_ID(resultCdRef)
                        id, resultCdVer = definition.get_ID(resultCdVer)
                        ts0_sheet.cell(row=x, column=9).value = resultCd                   
                        ts0_sheet.cell(row=x, column=10).value = resultCdRef  
                        ts0_sheet.cell(row=x, column=11).value = resultCdVer
            # filling TS Parameters sheet
            ts_sheet.cell(row=i, column=7).value = result2
            ts_sheet.cell(row=i, column=8).value = " "   
            if result2 == " ": ts_sheet.cell(row=i, column=8).value = nfValue	
            ts_sheet.cell(row=i, column=9).value = resultCd  
            ts_sheet.cell(row=i, column=10).value = resultCdRef
            ts_sheet.cell(row=i, column=11).value = resultCdVer
        file.close
