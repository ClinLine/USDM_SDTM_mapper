

import xml.etree.cElementTree as ET
import openpyxl
#import xls2define
import odmlib
#from odmlib.define_2_0 import model as DEFINE
import odmlib.ns_registry as NS


def Create_Define(wb,ta_var):
    _create_itemgroupdef_object(wb)
    content = _create_itemgroupdef_object(wb)
    _IGOID = "MDV1"
    IGVersion = "3.2.1"
    ODM_PREFIX = "odm:"
    XLMS_PREFIX = "xlms:"
    
    root = ET.Element("ODM", FileType="Snapshot", ODMVersion="1.3.2", CreationDateTime="2024-06-01T12:00:00", xmlns="http://www.cdisc.org/ns/odm/v1.3", XLMS_def="http://www.cdisc.org/ns/def/v2.1")
    study = ET.SubElement(root, "Study", OID="STUDY1")

    metaDataVersion = ET.SubElement(study, "MetaDataVersion", OID=_IGOID, Name="Metadata Version 1", Description="Metadata Version 1")
    ET.SubElement(metaDataVersion, "Standards", Name="SDTMIG", Type="IG", Version=IGVersion, CommentOID="COM.STD1")
    for domain in range(0, len(content)):
        if content[domain][0] is not None:
            # "OID","Name", "Repeating","Domain", "SASDatasetName", "IsReferenceData", "Purpose", "Class", "Structure","ArchiveLocationID","Description"
            domainElement = ET.SubElement(metaDataVersion, "ItemGroupDef", OID=content[domain][0], Name=content[domain][1], Domain=content[domain][3], Purpose=content[domain][6],SasDatasetName=content[domain][4], Repeating=content[domain][2], IsReferenceData=content[domain][5], Structure=content[domain][8], ArchiveLocation=content[domain][9], StandardOID=_IGOID)
            Description=ET.SubElement(domainElement, "Description")
            ET.SubElement(Description, "TranslatedText", xml_lang="en").text = content[domain][10]
            if content[domain][1] == "TA":
                AddTARef(ta_var, ET, domainElement)
    #         # ET.SubElement(domainElement, "def:Class", Name="TRIAL DESIGN")
    AddMethods(ta_var,ET,metaDataVersion)

    tree=ET.ElementTree(root)
    tree.write("Output/Define.xml")

def AddTARef(ta_var,ET,domainElement):
    for var in ta_var:
        if ta_var[var][0] is not None:
           # print(ta_var[var])
            if ta_var[var][0] == "STUDYID":
                itemOID = "IT.STUDYID"
            else:
                itemOID = "IT.TA." + ta_var[var][0]
            if ta_var[var][4] == "Req":
                isMandatory = "Yes"
            else:            
                isMandatory = "No"
            if ta_var[var][0] == "TAETORD":
                methodOID = "MT.TAETORD"
                ET.SubElement(domainElement, "ItemRef", ItemOID=itemOID, Mandatory=isMandatory, OrderNumber=str(var), MethodOID=methodOID)  
            else:
                ET.SubElement(domainElement, "ItemRef", ItemOID=itemOID, Mandatory=isMandatory, OrderNumber=str(var))  

def AddMethods(ta_var,ET,metaDataVersion):
    for var in ta_var:
        if ta_var[var][0] is not None:
            if ta_var[var][0] == "TAETORD":
                methodOID = "MT.TAETORD"
                Method=ET.SubElement(metaDataVersion, "MethodDef", OID=methodOID, Name="TAETORD Derivation Method", Type="Computation")
                Description=ET.SubElement(Method, "Description")
                ET.SubElement(Description, "TranslatedText", xml_lang="en").text = "Sequential number identifying the order of epochs within an arm. Based on the previous and next epoch start date and time. The first epoch of an arm is assigned a TAETORD of 1."

def _create_itemgroupdef_object(wb):
    """
    use the values from the Dataset worksheet row to create a ItemGroupDef odmlib object
    :param row: Datasets worksheet row values as a dictionary
    :return: odmlib ItemGroupDef object
    - Used code from https://github.com/swhume/odmlib_examples/blob/master/xls2define/Datasets.py
    """
    
    sh_D = wb['Domains']
    datasets_dict = {}
    datasets_dict[0]=["OID","Name", "Repeating","Domain", "SASDatasetName", "IsReferenceData", "Purpose", "Class", "Structure","ArchiveLocationID","Description"]
    for r in range(2, sh_D.max_row+1):
        m=r-1
        datasets_dict[m]=[]
        #for i in range(1, sh_D.max_column+1):
        #    datasets_dict [m].append(sh_D.cell(row=r, column=i).value)
        datasets_dict[m].append("IG" + "." +  sh_D.cell(row=r, column=1).value) # oid
        datasets_dict[m].append(sh_D.cell(row=r, column=1).value) # name
        datasets_dict[m].append("Yes" if sh_D.cell(row=r, column=2).value == "Yes" else "No") # repeating
        datasets_dict[m].append(sh_D.cell(row=r, column=1).value) # domain
        datasets_dict[m].append(sh_D.cell(row=r, column=1).value)   # sasdatasetname
        datasets_dict[m].append(sh_D.cell(row=r, column=4).value)   # isreferencedata
        datasets_dict[m].append(sh_D.cell(row=r, column=3).value)   # purpose
        datasets_dict[m].append(sh_D.cell(row=r, column=5).value)   # class 
        datasets_dict[m].append(sh_D.cell(row=r, column=6).value)   # structure
        datasets_dict[m].append("LF" + "." +  sh_D.cell(row=r, column=1).value)   # archivelocationid
        datasets_dict[m].append(sh_D.cell(row=r, column=7).value)   # description
        #if row.get("Comment"):
        #attr["CommentOID"] = row["Comment"]
        # igd = DEFINE.ItemGroupDef(**attr)
        # tt = DEFINE.TranslatedText(_content=datasets_dict [m][6], lang=self.lang)
        # igd.Description = DEFINE.Description()
        # igd.Description.TranslatedText.append(tt)
        #DEFINE["ItemGroupDef"].append(igd)
    return datasets_dict