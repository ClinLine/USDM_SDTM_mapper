import jsonata
import json
import openpyxl
import pandas as pd

with open('ReCoPad.json', 'r') as file:
    data = json.load(file)

expr = jsonata.Jsonata("study.versions.studyDesigns.blindingSchema.standardCode.decode")
result = expr.evaluate(data)
print(result)

wb = openpyxl.load_workbook("Maps/sdtm_mapping_paths.xlsx")

# Access the 'TS Parameters' sheet
ts_sheet = wb['TS Parameters']

# Print the value in the first and seventh column of each row in the 'TS Parameters' sheet
for i in range(1, ts_sheet.max_row + 1):
    c1 = ts_sheet.cell(row=i, column=1).value
    print(c1)
    print(ts_sheet.cell(row=i, column=7).value)
    if ts_sheet.cell(row=i, column=7).value is not None:
        ts_sheet.cell(row=i, column=8).value = c1
    print(ts_sheet.cell(row=i, column=8).value)
wb.save("Maps/sdtm_mapping_paths.xlsx")