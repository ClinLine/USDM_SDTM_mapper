import jsonata
import json
import openpyxl
import pandas as pd
import definition

def Create_TA(wb, JsonInput):

    ta_sheet = wb['TA']
    r = ta_sheet.max_row
    rem = []
    ta_var = {}
    for i in range(2, ta_sheet.max_row + 1):
        j=i-1
        ta_var[j] = []
        #swap the rows and columns in the TI sheet
        for c in range(1, 6):
            ta_var[j].append(ta_sheet.cell(row=i, column=c).value)
        ta_sheet.cell(row=1, column=j).value = ta_var[j][0]    # add variable name as column header (for excel output)
        if ta_var[j][0] == "STUDYID":
            StudyIdCodeSnip = ta_sheet.cell(row=i, column=7).value
            StudyIDColumn = j
        elif ta_var[j][0] == "DOMAIN":
            DomainResult =  ta_sheet.cell(row=i, column=8).value
            DomainColumn = j
        elif ta_var[j][0] == "ARMCD":
            ArmNameCodeSnip = ta_sheet.cell(row=i, column=7).value
            ArmNameColumn = j
        elif ta_var[j][0] == "ARM":
            ArmDescriptionCodeSnip = ta_sheet.cell(row=i, column=7).value
            ArmDescriptionColumn = j
        elif ta_var[j][0] == "TAETORD":
            TaetordCodeSnip = ta_sheet.cell(row=i, column=7).value
            TaetordColumn = j
        elif ta_var[j][0] == "ETCD":
            EtcdCodeSnip = ta_sheet.cell(row=i, column=7).value
            EtcdColumn = j
        elif ta_var[j][0] == "ELEMENT":
            ElementCodeSnip = ta_sheet.cell(row=i, column=7).value
            ElementColumn = j
        elif ta_var[j][0] == "EPOCH":
            EpochCodeSnip = ta_sheet.cell(row=i, column=7).value
            EpochColumn = j
        elif ta_var[j][0] == "TABRANCH":
            BranchCodeSnip = ta_sheet.cell(row=i, column=7).value
            BranchColumn = j
        elif ta_var[j][0] == "TATRANS":
            TatransCodeSnip = ta_sheet.cell(row=i, column=7).value
            TatransColumn = j
        else:
            rem.append(j)
    
    for i in range(r, ta_sheet.max_column):
        for j in range(1, r):
            ta_sheet.cell(row=j, column=i).value = ""
    # Print the value in the first and seventh column of each row in the 'TS Parameters' sheet
    with open(JsonInput, 'r') as file:
        data=json.load (file)
        studyId=definition.Parse_jsonata(codeSnip=StudyIdCodeSnip,data=data)
        taetord=definition.Parse_jsonata(codeSnip=TaetordCodeSnip,data=data)        
        x=1
        resultArm=definition.Parse_jsonata(codeSnip=ArmDescriptionCodeSnip,data=data)
        resultArmName=definition.Parse_jsonata(codeSnip=ArmNameCodeSnip,data=data)
        epoch = definition.Parse_jsonata(codeSnip=EpochCodeSnip,data=data)
        resultArm2 = []
        resultArmName2 = []
        definition.string_to_list(resultArm, resultArm2)  # convert the string to a list
        definition.string_to_list(resultArmName, resultArmName2)  # convert the string to a list
        resultEl=definition.Parse_jsonata(codeSnip=EtcdCodeSnip,data=data)
        resultElName=definition.Parse_jsonata(codeSnip=ElementCodeSnip,data=data)
        resultBranch=definition.Parse_jsonata(codeSnip=BranchCodeSnip,data=data)
        resultTrans=definition.Parse_jsonata(codeSnip=TatransCodeSnip,data=data)
        resultEl2 = []
        resultElName2 = []
        resultElArm1 = []
        resultFromTrans = []
        resultToTrans = []
        epoch2 = []
        epochArm = []
        resultElArm2 = []
        taetord2 = []
        taetordArm = []
        definition.string_to_nested_list(taetord, taetordArm, taetord2)
        definition.string_to_nested_list(epoch, epochArm, epoch2)  # convert the string to a nested list
        definition.string_to_nested_list(resultEl, resultElArm1, resultEl2)  # convert the string    
        definition.string_to_nested_list(resultElName, resultElArm2, resultElName2)  # convert the string
        definition.string_to_nested_list(resultTrans, resultFromTrans, resultToTrans)  # convert the string
        ArmNameId2 = {}
        ArmId2 = {}
        epochId2 = {}
        taetordOrderNext = {}
        taetordOrderNow = {}
        taetordArm3 = {}
        taetordArm4 = []
        taetordId2 = []
        rowIds = []
        for i in range(len(taetord2)):
            taetordEpoch, taetordNext = definition.get_ID(taetord2[i])  # extracting the ID from the string
            taetordArm2, taetordId = definition.get_ID(taetordArm[i])  # extracting the ID from the string   
            # note that we assume the same order of items in the resultElArm1 and resultEl2 lists, which should be the case if the input is correct. We also assume that the order of items in the taetordArm and taetord2 lists is the same.
            ElId, resultEl3 = definition.get_ID(resultEl2[i])  # extracting the ID from the string
            ElNameId, resultElName3 = definition.get_ID(resultElName2[i])  # extracting the ID from the string
            
            taetordOrderNext[taetordId] = taetordNext  # store the next order in a dictionary with the epoch as key
            taetordOrderNow[taetordId] = taetordEpoch  # store the current order in a dictionary with the epoch as key
            taetordArm3[taetordId] = taetordArm2  # store the arm code in a dictionary with the ID as key
            taetordId2.append(taetordId)  # store the ID in a list
            if taetordArm2 not in taetordArm4:  # check if the arm code is already in the list
                taetordArm4.append(taetordArm2)

            rowId = {
                     "CellId": taetordId,
                     "ArmId": taetordArm2,
                     "EpochId": taetordEpoch,
                     "NextEPochId": taetordNext,
                     "taeord": 0, # placeholder, will be filled later
                     "ElCode": resultEl3,
                     "ElName": resultElName3,
                     "tabranch": " " # placeholder, will be filled later 
                 }
            rowIds.append(rowId)
          
        rowIds = sort_row_ids_by_epoch(rowIds)
        if resultBranch == "": resultBranch = "Allocated" #If randomized or stratified not specified in characteristics.
        rowIds = AddTABranches(rowIds,resultBranch +" to")


        for item in epoch2:
            epochId, epoch3 = definition.get_ID(item)  # extracting the ID from the string
            epochId2[epochId] = epoch3  # store the epoch code in a dictionary with the ID as key
        for item in resultArmName2:  
            ArmNameId, resultArmName3 = definition.get_ID(item)
            ArmNameId2[ArmNameId] = resultArmName3  # store the arm name in a dictionary with the ID as key
        for item in resultArm2:
            ArmId, resultArm3 = definition.get_ID(item)  # extracting the ID from the string
            ArmId2[ArmId] = resultArm3  # store the arm code in a dictionary with the ID as key
               
        # add trans information if applicable - epoch id needs to be resolved to epoch name still.
        for i in range(len(resultFromTrans)):
            fromId, resultFromTrans2 = definition.get_ID(resultFromTrans[i])  # extracting the ID from the string
            toId, resultToTrans2 = definition.get_ID(resultToTrans[i])  # extracting the ID from the string
            if fromId != resultToTrans2: # check if there is actually a transformation to another epoch
                for row in rowIds:
                    if row["EpochId"] == fromId:
                        row["tatrans"] =  f"{resultFromTrans2}: transition to " + resultToTrans2  # add the transformation information to the relevant rows
                               
        n = 0
        for item in rowIds:
            x += 1
            ta_sheet.cell(row=x, column=TaetordColumn).value = item["taeord"]
            ta_sheet.cell(row=x, column=StudyIDColumn).value = studyId
            ta_sheet.cell(row=x, column=DomainColumn).value = DomainResult 
            ta_sheet.cell(row=x, column=ArmNameColumn).value = ArmNameId2[item["ArmId"]]
            ta_sheet.cell(row=x, column=ArmDescriptionColumn).value = ArmId2[item["ArmId"]]
            ta_sheet.cell(row=x, column=EtcdColumn).value = item["ElCode"]
            ta_sheet.cell(row=x, column=ElementColumn).value = item["ElName"]
            ta_sheet.cell(row=x, column=EpochColumn).value = epochId2[item["CellId"]]
            ta_sheet.cell(row=x, column=BranchColumn).value = item["tabranch"]
            ta_sheet.cell(row=x, column=TatransColumn).value = item.get("tatrans", "")
            n += 1
        for j in rem:
            for i in range(2, x + 1):
                ta_sheet.cell(row=i, column=j).value = ""
    return ta_var

def AddTABranches(row_ids, prefix=""):
    if not row_ids:
            return row_ids
    # group the rows by their order number (taeord)
    OrderElements = {}
    for row in row_ids:
        ord_no = row["taeord"]
        if ord_no not in OrderElements:
            OrderElements[ord_no] = []
        OrderElements[ord_no].append(row)
    
    BranchOrdNo=999
    # find the order number where the epoch branches start (where there are multiple elements with different code with the same order number) 
    # and define the elements which need to be shown in the branch information
    ShowOrd = []
    for ord_no, el_rows in OrderElements.items():
        ElCode=""
        for row in el_rows:
            if ElCode=="": ElCode = row["ElCode"]
            if row["taeord"]==1: ElCode=row["ElCode"] 
            elif row["ElCode"] != ElCode:
                if BranchOrdNo == 999: 
                    BranchOrdNo = row["taeord"]-1
                if row["taeord"] not in ShowOrd:
                    ShowOrd.append(row["taeord"])

    BranchText = {}
    # define branch text per arm
    for ord_no, el_rows in OrderElements.items():
        for row in el_rows:
            arm_id = row.get("ArmId")
            if row["taeord"] in ShowOrd:
                el_name = row.get("ElName", "")
                if arm_id not in BranchText or BranchText[arm_id] == "":
                    BranchText[arm_id] = el_name
                elif el_name != BranchText[arm_id]:
                    BranchText[arm_id] = f"{BranchText[arm_id]} {el_name}" if el_name else BranchText[arm_id]

    for row in row_ids:
        if row["taeord"] == BranchOrdNo:
            row["tabranch"] = prefix + " " + BranchText.get(row["ArmId"], "")

    return row_ids
            
    
        
def sort_row_ids_by_epoch(row_ids):
            """Sort rowIds following the epoch chain from start to end, within each ArmId"""
            if not row_ids:
                return row_ids
            
            # Group rows by ArmId
            arm_groups = {}
            for row in row_ids:
                arm_id = row["ArmId"]
                if arm_id not in arm_groups:
                    arm_groups[arm_id] = []
                arm_groups[arm_id].append(row)
            
            sorted_rows = []
                        
            # Process each ArmId group
            for arm_id, arm_rows in arm_groups.items():
                # Create a mapping of EpochId -> list of rowIds for this arm
                epoch_map = {}
                for row in arm_rows:
                    epoch_id = row["EpochId"]
                    if epoch_id not in epoch_map:
                        epoch_map[epoch_id] = []
                    epoch_map[epoch_id].append(row)
                
                # Find starting epochs (where no other epoch points to them as NextEPochId)
                next_epochs = {row["NextEPochId"] for row in arm_rows if row["NextEPochId"] != "None"}
                start_epochs = [row["EpochId"] for row in arm_rows if row["EpochId"] not in next_epochs]
                
                # Build the epoch chain order for this arm
                arm_sorted = []
                visited_epochs = set()
                
                def traverse_epoch_chain(epoch_id):
                    if epoch_id in visited_epochs or epoch_id == "None" or epoch_id not in epoch_map:
                        return
                    visited_epochs.add(epoch_id)
                    arm_sorted.extend(epoch_map[epoch_id])
                    
                    # Find the next epoch in the chain
                    for row in arm_rows:
                        if row["EpochId"] == epoch_id and row["NextEPochId"] != "None":
                            traverse_epoch_chain(row["NextEPochId"])
                            break
                
                # Traverse from each start epoch
                for start_epoch in start_epochs:
                    traverse_epoch_chain(start_epoch)
                
                # Add any remaining rows that weren't reached for this arm
                for row in arm_rows:
                    if not any(r["CellId"] == row["CellId"] for r in arm_sorted):
                        arm_sorted.append(row)
                
                # Assign order numbers (taeord) within this arm
                for order_num, row in enumerate(arm_sorted, start=1):
                    row["taeord"] = order_num
                
                sorted_rows.extend(arm_sorted)
            
            return sorted_rows
        
   